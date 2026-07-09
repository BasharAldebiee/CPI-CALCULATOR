"""
ABS 6401.0 CPI Data Parser
Handles the real ABS Time Series Workbook structure.

Supported files:
  640101.xlsx  — Monthly CPI Indicator (Table 1), all 9 cities, ~68% basket
  6401017.xlsx — Quarterly CPI All Groups (Table 17), all 9 cities, 100% basket

Table 1 Data1 layout (27 data columns):
  Row 1:  Series description headers
  Rows 2-10: Metadata
  Row 11+: Date + values
  Cols 1-9:   Index Numbers (Australia first, then 8 capitals)
  Cols 10-18: YoY % Change
  Cols 19-27: MoM % Change

Table 17 Data1 layout (18 data columns):
  Row 1:  Series description headers
  Rows 2-10: Metadata
  Row 11+: Date + values
  Cols 1-9:   Index Numbers (Sydney first, Australia last)
  Cols 10-18: QoQ % Change (same city order)
  No YoY column — computed from index numbers.
"""

import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import requests

CITIES = ['Australia', 'Sydney', 'Melbourne', 'Brisbane', 'Adelaide',
          'Perth', 'Hobart', 'Darwin', 'Canberra']

CITY_COL_MAP = {  # 0-indexed column in Data1 sheet
    'index': list(range(1, 10)),   # cols 1-9
    'yoy':   list(range(10, 19)),  # cols 10-18
    'mom':   list(range(19, 28)),  # cols 19-27
}

ABS_URLS = {
    'monthly':   'https://www.abs.gov.au/statistics/economy/price-indexes-and-inflation/consumer-price-index-australia/may-2026/640101.xlsx',
    # Table 17 — full-basket quarterly CPI, updated each quarterly release (Mar/Jun/Sep/Dec)
    'quarterly': 'https://www.abs.gov.au/statistics/economy/price-indexes-and-inflation/consumer-price-index-australia/mar-2026/6401017.xlsx',
}

# Table 17 city column order (Sydney first, Australia last) — differs from CITIES
_T17_CITY_ORDER = ['Sydney', 'Melbourne', 'Brisbane', 'Adelaide', 'Perth',
                   'Hobart', 'Darwin', 'Canberra', 'Australia']


def fetch_from_url(url: str) -> BytesIO:
    """Download an ABS Excel file from URL and return as BytesIO."""
    headers = {
        'User-Agent': 'Mozilla/5.0 (compatible; ABS-CPI-Tool/3.0)',
        'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*',
    }
    response = requests.get(url, headers=headers, timeout=60)
    response.raise_for_status()
    return BytesIO(response.content)


def load_abs_url(series: str) -> pd.DataFrame:
    """
    Fetch and parse ABS CPI data directly from the ABS website.
    series: 'monthly'   → Table 1  (640101.xlsx,  monthly indicator, 68% basket)
            'quarterly' → Table 17 (6401017.xlsx, full-basket quarterly CPI)
    """
    file_obj = fetch_from_url(ABS_URLS[series])
    if series == 'quarterly':
        return _parse_table17(file_obj)
    return load_abs_file(file_obj)


def _detect_frequency(ws_rows) -> str:
    """Detect Monthly vs Quarterly from the Frequency metadata row (row 5)."""
    freq_row = ws_rows[4]  # row index 4 = Excel row 5
    for val in freq_row[1:]:
        if isinstance(val, str):
            v = val.strip().lower()
            if 'quarter' in v:
                return 'Quarterly'
            if 'month' in v:
                return 'Monthly'
    # Fall back: check date spacing
    dates = [r[0] for r in ws_rows[10:13] if r[0] is not None]
    if len(dates) >= 2:
        delta = (pd.Timestamp(dates[1]) - pd.Timestamp(dates[0])).days
        return 'Quarterly' if delta > 60 else 'Monthly'
    return 'Monthly'


def load_abs_file(filepath_or_bytes) -> pd.DataFrame:
    """
    Parse ABS 6401.0 Table 1 or Table 6 Excel file (monthly or quarterly).
    Accepts: URL string, bytes, BytesIO, or file-like object.

    Returns clean DataFrame with columns:
      Date, Period, Frequency, {City}_Index, {City}_YoY, {City}_MoM
      for each of 9 cities.
    """
    # ── Handle URL string ──────────────────────────────────────────────────
    if isinstance(filepath_or_bytes, str):
        if filepath_or_bytes.startswith('http'):
            filepath_or_bytes = fetch_from_url(filepath_or_bytes)
        else:
            # Local file path
            with open(filepath_or_bytes, 'rb') as f:
                filepath_or_bytes = BytesIO(f.read())

    # ── Handle raw bytes ───────────────────────────────────────────────────
    elif isinstance(filepath_or_bytes, bytes):
        filepath_or_bytes = BytesIO(filepath_or_bytes)

    # At this point filepath_or_bytes is BytesIO or a file-like object
    wb = load_workbook(filepath_or_bytes, read_only=True, data_only=True)

    if 'Data1' not in wb.sheetnames:
        raise ValueError(
            "Sheet 'Data1' not found. "
            "Please use an ABS 6401.0 Table 1 or Table 6 Time Series workbook."
        )

    ws = wb['Data1']
    rows = list(ws.iter_rows(values_only=True))

    frequency = _detect_frequency(rows)

    # Data starts at row index 10 (row 11 in Excel)
    records = []
    for r in rows[10:]:
        dt = r[0]
        if dt is None:
            continue
        record = {'Date': pd.Timestamp(dt), 'Frequency': frequency}
        for i, city in enumerate(CITIES):
            record[f'{city}_Index'] = _safe_float(r[CITY_COL_MAP['index'][i]])
            record[f'{city}_YoY']   = _safe_float(r[CITY_COL_MAP['yoy'][i]])
            record[f'{city}_MoM']   = _safe_float(r[CITY_COL_MAP['mom'][i]])
        records.append(record)

    df = pd.DataFrame(records).sort_values('Date').reset_index(drop=True)

    # Format period label
    if frequency == 'Quarterly':
        df['Period'] = df['Date'].apply(_quarter_label)
    else:
        df['Period'] = df['Date'].dt.strftime('%b-%Y')

    return df


def _parse_table17(file_obj) -> pd.DataFrame:
    """
    Parse Table 17 (6401017.xlsx): CPI Quarterly All Groups, full basket.

    Col layout (0-indexed after row 1 headers):
      Cols 1-9  → Index Numbers for Sydney, Melbourne, Brisbane, Adelaide,
                  Perth, Hobart, Darwin, Canberra, Australia
      Cols 10-18 → QoQ % change (same city order)
      YoY is not in the file; computed here as (current / same-Q-prior-year) - 1.
    """
    wb   = load_workbook(file_obj, read_only=True, data_only=True)
    ws   = wb['Data1']
    rows = list(ws.iter_rows(values_only=True))

    records = []
    for r in rows[10:]:          # data starts at Excel row 11
        dt = r[0]
        if dt is None:
            continue
        record = {'Date': pd.Timestamp(dt), 'Frequency': 'Quarterly'}
        for i, city in enumerate(_T17_CITY_ORDER):
            record[f'{city}_Index'] = _safe_float(r[i + 1])   # cols 1-9
            record[f'{city}_MoM']   = _safe_float(r[i + 10])  # cols 10-18
            record[f'{city}_YoY']   = None                     # filled below
        records.append(record)

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records).sort_values('Date').reset_index(drop=True)
    df['Period'] = df['Date'].apply(_quarter_label)

    # Compute YoY: current quarter index vs same quarter in prior year
    p2i = {row['Period']: i for i, row in df.iterrows()}
    for i, row in df.iterrows():
        period = row['Period']
        prev_q = f"Q{period[1]}-{int(period[3:]) - 1}"
        if prev_q not in p2i:
            continue
        j = p2i[prev_q]
        for city in _T17_CITY_ORDER:
            curr = row[f'{city}_Index']
            prev = df.at[j, f'{city}_Index']
            if curr is not None and prev is not None and prev != 0:
                df.at[i, f'{city}_YoY'] = round((curr - prev) / prev * 100, 2)

    return df


def _quarter_label(dt: pd.Timestamp) -> str:
    """Convert a date to ABS quarter label e.g. Q1-2026 (Jan-Mar)."""
    q = (dt.month - 1) // 3 + 1
    return f"Q{q}-{dt.year}"


def derive_quarterly_from_monthly(df: pd.DataFrame) -> pd.DataFrame:
    """
    Derive quarterly CPI matching ABS methodology (from April 2025 releases):
      - YoY:  quarter-end month vs same month prior year  (e.g. Mar-2026 vs Mar-2025)
              This matches the ABS headline quarterly CPI (e.g. 4.6% for Q1-2026).
      - QoQ:  quarterly AVERAGE vs previous quarter average
              (e.g. avg(Jan+Feb+Mar-2026) / avg(Oct+Nov+Dec-2025) - 1 = 1.4%)
    Only complete quarters (all 3 months present) are included.
    """
    df = df.copy()
    df['_q'] = df['Date'].apply(_quarter_label)

    # Build quarterly averages for QoQ (all cities)
    q_avgs: dict = {}
    for quarter, grp in df.groupby('_q'):
        if len(grp) >= 3:
            q_avgs[quarter] = {city: float(grp[f'{city}_Index'].mean())
                               for city in CITIES}

    if not q_avgs:
        return pd.DataFrame()

    sorted_q = sorted(q_avgs, key=lambda q: (int(q[3:]), int(q[1])))

    # Keep only quarter-end months (Mar/Jun/Sep/Dec) for complete quarters
    quarter_ends = {3, 6, 9, 12}
    df_q = df[df['Date'].dt.month.isin(quarter_ends)].copy()
    df_q['Period'] = df_q['Date'].apply(_quarter_label)
    df_q = df_q[df_q['Period'].isin(q_avgs)].sort_values('Date').reset_index(drop=True)
    df_q['Frequency'] = 'Quarterly'

    # Compute QoQ from quarterly averages; YoY comes directly from the file's column
    for i, row in df_q.iterrows():
        period = row['Period']
        qi = sorted_q.index(period) if period in sorted_q else -1
        for city in CITIES:
            if qi <= 0:
                df_q.at[i, f'{city}_MoM'] = None
            else:
                prev_q = sorted_q[qi - 1]
                curr_a = q_avgs[period][city]
                prev_a = q_avgs[prev_q][city]
                df_q.at[i, f'{city}_MoM'] = round((curr_a - prev_a) / prev_a * 100, 2)

    return df_q.reset_index(drop=True)


def _safe_float(val):
    try:
        return float(val) if val is not None else None
    except (TypeError, ValueError):
        return None


def inspect_abs_file(series: str = 'quarterly') -> dict:
    """
    Diagnostic: fetch a raw ABS file and return its actual column headers,
    frequency metadata, and first data row so you can verify the column mapping
    matches CITY_COL_MAP and CITIES ordering.

    Usage:
        from abs_cpi_parser import inspect_abs_file
        import pprint; pprint.pprint(inspect_abs_file('quarterly'))
    """
    url = ABS_URLS[series]
    file_obj = fetch_from_url(url)
    wb = load_workbook(file_obj, read_only=True, data_only=True)

    if 'Data1' not in wb.sheetnames:
        return {'error': "Sheet 'Data1' not found"}

    ws = wb['Data1']
    rows = list(ws.iter_rows(values_only=True))

    meta_labels = ['series_desc', 'unit', 'series_type', 'data_type',
                   'frequency', 'collection_month', 'series_start',
                   'series_end', 'num_obs', 'data_starts_row11']

    result = {
        'sheets':         wb.sheetnames,
        'detected_freq':  _detect_frequency(rows),
        'total_rows':     len(rows),
        'column_headers': list(rows[0]),           # row 1: series descriptions
        'num_columns':    len(rows[0]),
        'metadata': {},
        'first_data_row': list(rows[10]) if len(rows) > 10 else [],
        'last_data_row':  list(rows[-1]),
    }
    for i, label in enumerate(meta_labels):
        if i < len(rows):
            result['metadata'][label] = list(rows[i])

    return result


def get_city_df(df: pd.DataFrame, city: str) -> pd.DataFrame:
    """Extract single-city view with renamed columns."""
    cols = ['Date', 'Period', 'Frequency', f'{city}_Index', f'{city}_YoY', f'{city}_MoM']
    out = df[cols].copy()
    out.columns = ['Date', 'Period', 'Frequency', 'CPI_Index', 'YoY_Pct', 'MoM_Pct']
    return out


def calc_custom_change(df: pd.DataFrame, city: str, start_period: str, end_period: str) -> dict:
    """Calculate CPI movement between any two periods."""
    city_df = get_city_df(df, city)
    s = city_df[city_df['Period'] == start_period]['CPI_Index'].values
    e = city_df[city_df['Period'] == end_period]['CPI_Index'].values
    if len(s) == 0 or len(e) == 0:
        return {}
    start_val, end_val = float(s[0]), float(e[0])
    pct = ((end_val - start_val) / start_val) * 100
    return {
        'start_period': start_period,
        'end_period':   end_period,
        'start_val':    start_val,
        'end_val':      end_val,
        'movement':     end_val - start_val,
        'pct_change':   pct,
    }
