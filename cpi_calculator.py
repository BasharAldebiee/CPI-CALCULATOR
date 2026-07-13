"""
Australia Post — CPI Calculator
Data & Pricing Executive Tool  |  Version 3.1
Source: ABS 6401.0 Consumer Price Index, Australia (Monthly + Quarterly)
Base period: September 2025 = 100.00
Data fetched live from ABS website — no manual uploads required.
"""


import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

from abs_cpi_parser import load_abs_url, get_city_df, calc_custom_change, CITIES, derive_quarterly_from_monthly

# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CPI Calculator | Australia Post",
    page_icon="📮",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── STYLES ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500&family=IBM+Plex+Sans:wght@300;400;500;600;700&display=swap');

/* ── Design tokens ───────────────────────────────────────────────── */
:root {
  /* Brand + accents (brightened for legibility on dark surfaces) */
  --red:    #C8102E; --red-700:#9B0B23; --red-bright:#FF5A73;
  --blue:   #60A5FA; --amber: #FBBF24; --green: #34D399;
  /* Ink ramp — light text on dark */
  --ink-900:#F4F4F7; --ink-700:#CFCDD9; --ink-600:#A7A4B6; --ink-400:#6E6B7E;
  /* Surfaces */
  --border: #2B2836; --surface:#0C0A12; --card:#17151F; --card-2:#1F1C2B;
  --sidebar-bg:#100E17;
  --white:  #FFFFFF;   /* kept literal — used for the logo plate */
  /* Notice ink (readable on the tinted notice backgrounds) */
  --amber-ink:#FCD9A6; --blue-ink:#C7DCFF; --green-ink:#B7F0D2;
  /* Text scale */
  --t-xs:0.75rem; --t-sm:0.875rem; --t-base:1rem; --t-lg:1.25rem; --t-xl:1.75rem;
  /* Numeric (mono) ramp — one deliberate 3-step hierarchy for figures */
  --num-sm:1.1rem; --num-md:1.2rem; --num-lg:1.9rem;
  /* 4pt base spacing */
  --s-1:4px; --s-2:8px; --s-3:12px; --s-4:16px; --s-5:20px; --s-6:24px; --s-8:32px;
  /* Elevation & radius */
  --radius:10px; --radius-sm:6px;
  --shadow-sm:0 1px 2px rgba(0,0,0,0.30), 0 1px 3px rgba(0,0,0,0.40);
  --shadow-md:0 2px 6px rgba(0,0,0,0.40), 0 8px 24px rgba(0,0,0,0.45);
  --focus:0 0 0 2px var(--surface), 0 0 0 4px var(--blue);
}

/* Accessible keyboard focus — visible 2px ring for switch/keyboard users */
:focus-visible { outline:none; box-shadow:var(--focus); border-radius:var(--radius-sm); }

/* Enterprise chrome — hide only Streamlit branding, keep the ⋮ menu/toolbar usable */
footer { visibility:hidden; }
[data-testid="stDecoration"] { display:none; }
.block-container { padding-top:2rem; }

/* Respect users who disable motion */
@media (prefers-reduced-motion: reduce) {
  *, *::before, *::after { animation:none !important; transition:none !important; }
}

/* Theme robustness — white cards always keep dark ink, never inherit a light
   base text colour (prevents invisible text if the app is forced to dark). */
.kpi, .dual-header, .signal-box { color:var(--ink-900); }

html, body, [class*="css"] { font-family:'IBM Plex Sans',sans-serif; background:var(--surface); }

/* ── Header (compact toolbar) ────────────────────────────────────── */
.ap-header {
  background:
    radial-gradient(120% 160% at 88% 12%, rgba(255,255,255,0.12) 0%, transparent 46%),
    linear-gradient(135deg, #EB2D3C 0%, #DC1928 60%, #B01420 100%);
  padding:var(--s-5) var(--s-6);
  border-radius:var(--radius); margin-bottom:var(--s-6);
  box-shadow:var(--shadow-md);
  display:flex; align-items:center; gap:var(--s-4);
  overflow:hidden; position:relative;
}
.ap-header h1 {
  color:white; font-size:1.55rem; font-weight:700;
  margin:0; letter-spacing:-0.02em; text-shadow:0 1px 12px rgba(0,0,0,0.35);
}
.ap-header .sub {
  color:rgba(255,255,255,0.92); font-size:var(--t-xs);
  margin-top:3px; display:flex; align-items:center; gap:var(--s-2); flex-wrap:wrap;
}
.ap-logo { font-size:1.4rem; flex-shrink:0; }
/* Subtle white plate — keeps a dark/multi-colour logo separated from the red
   header. A clean white-on-transparent mark can skip this via .ap-logo-bare. */
.ap-logo-img {
  height:38px; width:auto; flex-shrink:0; display:block;
  background:var(--white); border-radius:var(--radius-sm);
  padding:6px 10px; box-shadow:var(--shadow-sm);
}
.ap-logo-img.ap-logo-bare {
  background:none; border-radius:0; padding:0; box-shadow:none;
}
.live-dot {
  display:inline-block; width:7px; height:7px; background:#4ADE80;
  border-radius:50%; animation:livepulse 2s ease-in-out infinite; flex-shrink:0;
}
@keyframes livepulse { 0%,100%{opacity:1} 50%{opacity:0.4} }

/* ── KPI cards ───────────────────────────────────────────────────── */
.kpi-wrap { display:flex; gap:var(--s-3); margin-bottom:var(--s-6); flex-wrap:wrap; align-items:stretch; }
.kpi {
  background:var(--card); border:1px solid var(--border); border-radius:var(--radius);
  box-shadow:var(--shadow-sm);
  padding:var(--s-3) var(--s-4); flex:1 1 128px; min-width:128px; overflow:hidden;
  display:flex; flex-direction:column; min-height:88px;
}
.kpi-label {
  font-size:var(--t-xs); font-weight:700; text-transform:uppercase;
  letter-spacing:0.06em; color:var(--ink-600); margin-bottom:var(--s-1);
  white-space:nowrap; overflow:hidden; text-overflow:ellipsis;
}
.kpi-val {
  font-family:'IBM Plex Mono',monospace; font-size:var(--num-md); font-weight:600;
  color:var(--ink-900); line-height:1.1; white-space:nowrap;
  overflow:hidden; text-overflow:ellipsis;
}
.kpi-sub {
  font-size:var(--t-xs); color:var(--ink-600); margin-top:auto; padding-top:var(--s-1);
  line-height:1.4; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;
}
/* Hero card — primary decision metric. Wins on width + size + colour. */
.kpi-hero {
  flex:1.8 1 200px; min-width:200px; border-left:4px solid var(--red-bright);
  background:rgba(200,16,46,0.16);
  padding:var(--s-3) var(--s-5);
}
.kpi-hero .kpi-label { color:var(--red-bright); }
.kpi-hero .kpi-val { font-size:var(--num-lg); }
.kpi-hero .kpi-sub { font-weight:600; }

.kpi-red   .kpi-val { color:var(--red-bright); }
.kpi-green .kpi-val { color:var(--green); }
.kpi-blue  .kpi-val { color:var(--blue); }
.kpi-compact .kpi-val   { font-size:var(--num-sm); }
.kpi-compact            { padding:10px 12px; }
.kpi-compact .kpi-label { font-size:0.70rem; }
.kpi-compact .kpi-sub   { font-size:0.70rem; }
.kpi-wrap-monthly   { border-top:3px solid var(--red);  padding-top:var(--s-2); }
.kpi-wrap-quarterly { border-top:3px solid var(--red-bright); padding-top:var(--s-2); }

/* ── Signal box ──────────────────────────────────────────────────── */
.signal-box {
  background:var(--card); border:1px solid var(--border); border-radius:var(--radius);
  box-shadow:var(--shadow-sm);
  padding:var(--s-4) var(--s-5); margin:var(--s-4) 0;
  display:flex; gap:var(--s-4); align-items:flex-start;
}
.signal-box-red   { border-left:4px solid var(--red); }
.signal-box-amber { border-left:4px solid var(--amber); }
.signal-box-green { border-left:4px solid var(--green); }
.signal-icon  { font-size:1.5rem; line-height:1; flex-shrink:0; }
.signal-title { font-weight:700; font-size:var(--t-base); color:var(--ink-900); margin-bottom:var(--s-1); }
.signal-desc  { font-size:var(--t-sm); color:var(--ink-600); line-height:1.5; }
.signal-formula {
  background:var(--card-2); border-radius:4px; padding:4px var(--s-2);
  font-family:'IBM Plex Mono',monospace; font-size:var(--t-xs);
  color:var(--ink-600); margin-top:var(--s-2); display:inline-block;
}

/* ── Pills & badges ──────────────────────────────────────────────── */
.series-pill {
  display:inline-block; padding:2px 8px; border-radius:20px;
  font-size:0.70rem; font-weight:700; letter-spacing:0.04em; vertical-align:middle;
}
.pill-monthly   { background:rgba(255,90,115,0.18); color:var(--red-bright); }
.pill-quarterly { background:rgba(255,90,115,0.18); color:var(--red-bright); }
.badge        { display:inline-block; background:rgba(255,90,115,0.18); color:var(--red-bright);
                font-size:0.70rem; font-weight:700; padding:2px 8px; border-radius:20px; vertical-align:middle; }
.badge-green  { background:rgba(52,211,153,0.18); color:var(--green); }
.badge-yellow { background:rgba(251,191,36,0.18); color:var(--amber); }

/* ── Section titles ──────────────────────────────────────────────── */
/* Neutral — red is reserved for signal/danger only */
.sec-title {
  font-size:var(--t-xs); font-weight:700; text-transform:uppercase;
  letter-spacing:0.10em; color:var(--ink-600);
  border-bottom:2px solid var(--border); display:block;
  padding-bottom:var(--s-1); margin-bottom:var(--s-3);
}
.sec-title-q {
  font-size:var(--t-xs); font-weight:700; text-transform:uppercase;
  letter-spacing:0.10em; color:var(--red-bright);
  border-bottom:2px solid var(--red-bright); display:block;
  padding-bottom:var(--s-1); margin-bottom:var(--s-3);
}
.sec-title-sidebar {
  font-size:0.68rem; font-weight:700; text-transform:uppercase;
  letter-spacing:0.10em; color:var(--ink-600); border-bottom:1px solid var(--border);
  display:block; padding-bottom:3px; margin-bottom:var(--s-2); margin-top:var(--s-5);
}

/* ── Context bar (snapshot summary strip) ────────────────────────── */
.dual-header {
  background:var(--card); border:1px solid var(--border); border-radius:var(--radius);
  box-shadow:var(--shadow-sm);
  padding:var(--s-4) var(--s-5); margin-bottom:var(--s-4);
  display:flex; gap:var(--s-4); align-items:center; flex-wrap:wrap; row-gap:var(--s-3);
}
.dual-header-item { display:flex; flex-direction:column; gap:3px; min-width:0; flex-shrink:0; }
.dual-header-item.signal-item { flex:1; min-width:0; }
.dual-header-label {
  font-size:0.68rem; font-weight:700; text-transform:uppercase;
  letter-spacing:0.08em; color:var(--ink-600); white-space:nowrap;
}
.dual-header-val {
  font-family:'IBM Plex Mono',monospace; font-size:var(--num-sm);
  font-weight:600; color:var(--ink-900); white-space:nowrap;
}
.dual-monthly   { color:var(--red); }
.dual-quarterly { color:var(--red-bright); }

/* ── Evidence panel ──────────────────────────────────────────────── */
.ev-heading {
  font-size:0.70rem; font-weight:700; text-transform:uppercase;
  letter-spacing:0.08em; color:var(--ink-600); margin-bottom:var(--s-2);
}
.ev-row   { display:flex; align-items:center; gap:var(--s-2); margin:var(--s-1) 0; }
.ev-step  {
  background:var(--red); color:white; font-size:0.65rem; font-weight:700;
  border-radius:50%; width:20px; height:20px; display:inline-flex;
  align-items:center; justify-content:center; flex-shrink:0;
}
.ev-lbl     { font-size:var(--t-sm); color:var(--ink-600); min-width:180px; }
.ev-val     { font-family:'IBM Plex Mono',monospace; font-size:var(--t-base); font-weight:600; color:var(--ink-900); }
.ev-formula {
  background:var(--card-2); border:1px solid var(--border); border-radius:6px;
  padding:var(--s-3) var(--s-4); font-family:'IBM Plex Mono',monospace;
  font-size:var(--t-sm); color:var(--ink-600); flex:1; min-width:220px; line-height:1.9;
}
.ev-source {
  font-size:var(--t-xs); color:var(--ink-600);
  border-top:1px solid var(--border); padding-top:var(--s-2); margin-top:var(--s-3);
}

/* ── Utility components ──────────────────────────────────────────── */
.period-notice {
  background:rgba(251,191,36,0.12); border-left:3px solid var(--amber); border-radius:0 6px 6px 0;
  padding:var(--s-2) var(--s-4); font-size:var(--t-sm); color:var(--amber-ink); margin:0 0 var(--s-4) 0;
}
.abs-note {
  background:rgba(96,165,250,0.12); border-left:4px solid var(--blue); padding:var(--s-2) var(--s-4);
  border-radius:0 6px 6px 0; font-size:var(--t-sm); color:var(--blue-ink); margin:var(--s-2) 0;
}
.data-status {
  background:rgba(52,211,153,0.12); border-left:4px solid var(--green); padding:var(--s-2) var(--s-4);
  border-radius:0 6px 6px 0; font-size:var(--t-sm); color:var(--green-ink); margin:var(--s-1) 0;
}

/* ── Download button ─────────────────────────────────────────────── */
.stDownloadButton > button {
  background:var(--red) !important; color:white !important;
  border:none !important; font-weight:600 !important;
  border-radius:6px !important; padding:var(--s-2) var(--s-4) !important;
  font-size:var(--t-sm) !important;
}
.stDownloadButton > button:hover { background:var(--red-700) !important; }

/* ── Themed data tables (token-driven → follow the theme toggle) ───── */
.themed-table { width:100%; overflow-x:auto; margin:var(--s-2) 0 var(--s-4); }
.themed-table table {
  width:100%; border-collapse:collapse; font-size:0.86rem;
  background:var(--card); border:1px solid var(--border);
  border-radius:var(--radius); overflow:hidden;
}
.themed-table thead th {
  background:var(--card-2); color:var(--ink-600);
  font-size:0.68rem; font-weight:700; text-transform:uppercase; letter-spacing:0.05em;
  padding:10px 14px; border-bottom:1px solid var(--border); text-align:right; white-space:nowrap;
}
.themed-table tbody td {
  padding:9px 14px; border-bottom:1px solid var(--border);
  color:var(--ink-900); text-align:right;
  font-family:'IBM Plex Mono',monospace;
}
.themed-table th:first-child, .themed-table td:first-child {
  text-align:left; font-family:'IBM Plex Sans',sans-serif; font-weight:600;
}
.themed-table tbody tr:last-child td { border-bottom:none; }
.themed-table tbody tr:hover td { background:var(--card-2); }

/* ── Sidebar ─────────────────────────────────────────────────────── */
section[data-testid="stSidebar"] { background:var(--sidebar-bg); border-right:1px solid var(--border); }
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"],
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p,
section[data-testid="stSidebar"] .stRadio label,
section[data-testid="stSidebar"] .stCheckbox label,
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span { color:var(--ink-900) !important; }
</style>
""", unsafe_allow_html=True)

# ─── THEME SYSTEM ─────────────────────────────────────────────────────────────
# Two full palettes. The active one drives (a) a small :root CSS override and
# (b) the Python-side colour globals used by Plotly and pandas stylers.
THEMES = {
    "Dark": {
        # CSS :root tokens
        "red-bright": "#FF5A73", "blue": "#60A5FA", "amber": "#FBBF24", "green": "#34D399",
        "ink-900": "#F4F4F7", "ink-700": "#CFCDD9", "ink-600": "#A7A4B6", "ink-400": "#6E6B7E",
        "border": "#2B2836", "surface": "#0C0A12", "card": "#17151F", "card-2": "#1F1C2B",
        "sidebar-bg": "#100E17",
        "amber-ink": "#FCD9A6", "blue-ink": "#C7DCFF", "green-ink": "#B7F0D2",
        "shadow-sm": "0 1px 2px rgba(0,0,0,0.30), 0 1px 3px rgba(0,0,0,0.40)",
        "shadow-md": "0 2px 6px rgba(0,0,0,0.40), 0 8px 24px rgba(0,0,0,0.45)",
        # Python-side (charts / tables / inline)
        "monthly": "#FF5A73", "quarterly": "#FF5A73",
        "chart_bg": "#15131C", "chart_grid": "#2A2733", "chart_font": "#CFCDD9", "chart_pill": "#1F1C2B",
        "pos": "#FF5A73", "neg": "#34D399",
        "dot_bg": "#2A2733", "dot_fg": "#CFCDD9",
        "dot_nd_bg": "#1A1723", "dot_nd_fg": "#6E6B7E", "dot_nd_border": "#3A3746",
        "pill_bg": "#1F1C2B", "pill_border": "#2B2836", "pill_fg": "#CFCDD9",
    },
    "Light": {
        "red-bright": "#C8102E", "blue": "#1D4ED8", "amber": "#D97706", "green": "#16A34A",
        "ink-900": "#111827", "ink-700": "#374151", "ink-600": "#4B5563", "ink-400": "#9CA3AF",
        "border": "#E5E7EB", "surface": "#F9FAFB", "card": "#FFFFFF", "card-2": "#F3F4F6",
        "sidebar-bg": "#FAFAFA",
        "amber-ink": "#92400E", "blue-ink": "#1E40AF", "green-ink": "#14532D",
        "shadow-sm": "0 1px 2px rgba(16,24,40,0.04), 0 1px 3px rgba(16,24,40,0.06)",
        "shadow-md": "0 2px 4px rgba(16,24,40,0.06), 0 4px 12px rgba(16,24,40,0.08)",
        "monthly": "#C8102E", "quarterly": "#C8102E",
        "chart_bg": "#FFFFFF", "chart_grid": "#E5E7EB", "chart_font": "#111827", "chart_pill": "#FFFFFF",
        "pos": "#C8102E", "neg": "#15803D",
        "dot_bg": "#F1F5F9", "dot_fg": "#374151",
        "dot_nd_bg": "#F8FAFC", "dot_nd_fg": "#94A3B8", "dot_nd_border": "#CBD5E1",
        "pill_bg": "#F1F5F9", "pill_border": "#E2E8F0", "pill_fg": "#374151",
    },
}

# Active Python-side colour globals — reassigned by apply_theme() each run.
MONTHLY_COLOR   = "#FF5A73"
QUARTERLY_COLOR = "#60A5FA"
CHART_BG   = "#15131C"
CHART_GRID = "#2A2733"
CHART_FONT = "#CFCDD9"
CHART_PILL = "#1F1C2B"
POS_COLOR  = "#FF5A73"
NEG_COLOR  = "#34D399"
DOT_BG = "#2A2733"; DOT_FG = "#CFCDD9"
DOT_ND_BG = "#1A1723"; DOT_ND_FG = "#6E6B7E"; DOT_ND_BORDER = "#3A3746"
PILL_BG = "#1F1C2B"; PILL_BORDER = "#2B2836"; PILL_FG = "#CFCDD9"


def apply_theme(theme_name: str):
    """Set Python colour globals and inject the :root CSS override for the theme."""
    global MONTHLY_COLOR, QUARTERLY_COLOR, CHART_BG, CHART_GRID, CHART_FONT, CHART_PILL
    global POS_COLOR, NEG_COLOR, DOT_BG, DOT_FG, DOT_ND_BG, DOT_ND_FG, DOT_ND_BORDER
    global PILL_BG, PILL_BORDER, PILL_FG
    t = THEMES.get(theme_name, THEMES["Dark"])
    MONTHLY_COLOR, QUARTERLY_COLOR = t["monthly"], t["quarterly"]
    CHART_BG, CHART_GRID, CHART_FONT, CHART_PILL = t["chart_bg"], t["chart_grid"], t["chart_font"], t["chart_pill"]
    POS_COLOR, NEG_COLOR = t["pos"], t["neg"]
    DOT_BG, DOT_FG = t["dot_bg"], t["dot_fg"]
    DOT_ND_BG, DOT_ND_FG, DOT_ND_BORDER = t["dot_nd_bg"], t["dot_nd_fg"], t["dot_nd_border"]
    PILL_BG, PILL_BORDER, PILL_FG = t["pill_bg"], t["pill_border"], t["pill_fg"]

    root_vars = ";".join(
        f"--{k}:{t[k]}" for k in (
            "red-bright", "blue", "amber", "green",
            "ink-900", "ink-700", "ink-600", "ink-400",
            "border", "surface", "card", "card-2", "sidebar-bg",
            "amber-ink", "blue-ink", "green-ink", "shadow-sm", "shadow-md",
        )
    )
    # Force Streamlit's own native widgets (dropdowns, inputs, radios, checkboxes,
    # expanders) to follow the active theme — their internals are otherwise locked
    # to config.toml's base and won't flip with the toggle.
    bg, card, ink, border, ink6 = t["surface"], t["card"], t["ink-900"], t["border"], t["ink-600"]
    card2 = t["card-2"]
    widget_css = (
        # App + main canvas
        f".stApp,[data-testid=\"stAppViewContainer\"],[data-testid=\"stMain\"]{{background:{bg} !important;}}"
        # Selectbox / dropdown closed control
        f"[data-baseweb=\"select\"]>div{{background:{card} !important;border-color:{border} !important;}}"
        f"[data-baseweb=\"select\"] *{{color:{ink} !important;}}"
        f"[data-baseweb=\"select\"] svg{{fill:{ink} !important;}}"
        # Open dropdown menu — rendered in a portal at document root; theme every
        # nested container, list and option so nothing stays on the base surface.
        f"div[data-baseweb=\"popover\"] div,div[data-baseweb=\"popover\"] ul,"
        f"div[data-baseweb=\"popover\"] li,ul[role=\"listbox\"],[data-baseweb=\"menu\"]"
        f"{{background-color:{card} !important;}}"
        f"div[data-baseweb=\"popover\"] li,div[data-baseweb=\"popover\"] li *,"
        f"[role=\"option\"],[role=\"option\"] *{{color:{ink} !important;}}"
        f"div[data-baseweb=\"popover\"] li:hover,"
        f"div[data-baseweb=\"popover\"] li[aria-selected=\"true\"]{{background-color:{card2} !important;}}"
        # Text / number inputs
        f"[data-baseweb=\"input\"]>div,[data-baseweb=\"base-input\"]{{background:{card} !important;border-color:{border} !important;}}"
        f"[data-baseweb=\"input\"] input,[data-baseweb=\"base-input\"] input{{background:{card} !important;color:{ink} !important;}}"
        # Expander header
        f"[data-testid=\"stExpander\"] summary,[data-testid=\"stExpander\"] details{{background:{card} !important;color:{ink} !important;}}"
        f"[data-testid=\"stExpander\"] summary *{{color:{ink} !important;}}"
    )
    st.markdown(
        f"<style>:root{{{root_vars}}}{widget_css}</style>",
        unsafe_allow_html=True,
    )


def themed_plotly(fig, **kwargs):
    """Render a Plotly figure with axis tick labels + titles forced to the theme
    colour and bold. Plotly's template overrides layout.font on axes, so tick
    labels must be set explicitly per-axis — otherwise they render a faint grey."""
    fig.update_xaxes(tickfont=dict(color=CHART_FONT, weight="bold"),
                     title_font=dict(color=CHART_FONT, weight="bold"))
    fig.update_yaxes(tickfont=dict(color=CHART_FONT, weight="bold"),
                     title_font=dict(color=CHART_FONT, weight="bold"))
    st.plotly_chart(fig, **kwargs)


def render_themed_table(obj):
    """Render a pandas Styler (or DataFrame) as a token-driven HTML table that
    follows the theme toggle. Replaces st.dataframe, whose canvas-drawn cells
    cannot be recoloured by runtime CSS."""
    if hasattr(obj, "hide"):          # pandas Styler
        html = obj.hide(axis="index").to_html()
    else:                             # plain DataFrame
        html = obj.to_html(index=False, border=0)
    st.markdown(f'<div class="themed-table">{html}</div>', unsafe_allow_html=True)

# ─── DATA LOAD ────────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False, ttl=3600)
def fetch_monthly() -> pd.DataFrame | None:
    """Fetch monthly CPI data from ABS (Table 1, 640101.xlsx) — cached 1 hour."""
    try:
        return load_abs_url('monthly')
    except Exception as e:
        st.error(f"Could not fetch monthly data from ABS: {e}")
        return None


@st.cache_data(show_spinner=False, ttl=3600)
def fetch_quarterly() -> pd.DataFrame | None:
    """Fetch quarterly CPI data from ABS (Table 17, 6401017.xlsx) — cached 1 hour."""
    try:
        return load_abs_url('quarterly')
    except Exception as e:
        st.error(f"Could not fetch quarterly data from ABS (Table 17): {e}")
        return None




# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────

def build_excel_report(city_df, city, start_period, end_period, calc, frequency="Monthly"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CPI Report"

    RED, BLUE       = "C8102E", "1D4ED8"
    LIGHT_RED       = "FCE8EC"
    LIGHT_BLUE      = "DBEAFE"
    GREY, WHITE     = "F5F5F5", "FFFFFF"
    DARK            = "1A1A1A"
    HDR_COLOR       = RED if frequency == "Monthly" else BLUE
    LIGHT_COLOR     = LIGHT_RED if frequency == "Monthly" else LIGHT_BLUE
    thin = Side(style="thin", color="E0E0E0")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(cell, value, bg=None, fg="FFFFFF", sz=11, bold=True):
        cell.value = value
        cell.font = Font(name="Calibri", bold=bold, size=sz, color=fg)
        cell.fill = PatternFill("solid", fgColor=bg or HDR_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr

    ws.merge_cells("A1:G1")
    hdr_cell(ws["A1"], f"Australia Post — CPI Analysis Report ({frequency})", sz=14)
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Region: {city}   |   Period: {start_period} → {end_period}   |   ABS 6401.0  (Base: Sep-2025 = 100)"
    ws["A2"].font = Font(name="Calibri", size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    kpis = [
        ("Start Index",  f"{calc['start_val']:.2f}",  "D"),
        ("End Index",    f"{calc['end_val']:.2f}",    "E"),
        ("Movement",     f"{calc['movement']:+.2f} pts", "F"),
        ("% Change",     f"{calc['pct_change']:+.2f}%",  "G"),
    ]
    for label, value, col in kpis:
        ws[f"{col}4"] = label
        ws[f"{col}4"].font = Font(name="Calibri", bold=True, size=8, color="888888")
        ws[f"{col}4"].alignment = Alignment(horizontal="center")
        ws[f"{col}5"] = value
        ws[f"{col}5"].font = Font(name="Calibri", bold=True, size=16,
                                  color=HDR_COLOR if "%" in value or "pts" in value else DARK)
        ws[f"{col}5"].fill = PatternFill("solid", fgColor=LIGHT_COLOR if "%" in value else GREY)
        ws[f"{col}5"].alignment = Alignment(horizontal="center")
        ws[f"{col}5"].border = bdr
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 28

    pct = calc['pct_change']
    signal = "🔴 Strong Review Recommended" if abs(pct) >= 4 else ("🟡 Review Warranted" if abs(pct) >= 2 else "🟢 Stable — Monitor Quarterly")
    ws.merge_cells("D6:G6")
    ws["D6"] = f"Pricing Signal: {signal}"
    ws["D6"].font = Font(name="Calibri", bold=True, size=10, color=HDR_COLOR)
    ws["D6"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[6].height = 22

    period_header = "Period"
    pop_label = "QoQ Change (%)" if frequency == "Quarterly" else "MoM Change (%)"
    headers = [period_header, "CPI Index", pop_label, "YoY Change (%)", "vs Start (%)", "Signal"]
    for col_i, h in enumerate(headers, 1):
        c = ws.cell(row=8, column=col_i, value=h)
        hdr_cell(c, h, sz=9)
    ws.row_dimensions[8].height = 20

    start_date = city_df[city_df['Period'] == start_period]['Date'].values[0]
    end_date   = city_df[city_df['Period'] == end_period]['Date'].values[0]
    sub = city_df[
        (city_df['Date'] >= start_date) & (city_df['Date'] <= end_date)
    ].copy()
    sub['vs_start'] = ((sub['CPI_Index'] - calc['start_val']) / calc['start_val'] * 100).round(2)

    for row_i, (_, row) in enumerate(sub.iterrows(), 9):
        fill_c = "F9F9F9" if row_i % 2 == 0 else WHITE
        mom_val = row['MoM_Pct']
        yoy_val = row['YoY_Pct']
        vs_val  = row['vs_start']
        row_signal = "▲ Rising" if (mom_val or 0) > 0.5 else ("▼ Easing" if (mom_val or 0) < 0 else "→ Flat")
        vals = [row['Period'], row['CPI_Index'], mom_val, yoy_val, vs_val, row_signal]
        for col_i, v in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=col_i, value=v)
            c.font = Font(name="Calibri", size=9)
            c.alignment = Alignment(horizontal="center")
            c.fill = PatternFill("solid", fgColor=fill_c)
            c.border = bdr
            if col_i in (3, 4, 5) and isinstance(v, (int, float)) and v is not None:
                c.number_format = '+0.00;-0.00;0.00'
        ws.row_dimensions[row_i].height = 15

    for col, w in zip("ABCDEFG", [12, 11, 15, 14, 13, 12, 14]):
        ws.column_dimensions[get_column_letter(ord(col)-64)].width = w

    ws2 = wb.create_sheet("Raw Data")
    pop_col = "QoQ % Change" if frequency == "Quarterly" else "MoM % Change"
    raw_headers = ["Period", "CPI Index", pop_col, "YoY % Change"]
    for i, h in enumerate(raw_headers, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri")
        c.fill = PatternFill("solid", fgColor=HDR_COLOR)
        c.alignment = Alignment(horizontal="center")
    for row_i, (_, row) in enumerate(sub.iterrows(), 2):
        ws2.cell(row=row_i, column=1, value=row['Period'])
        ws2.cell(row=row_i, column=2, value=row['CPI_Index'])
        ws2.cell(row=row_i, column=3, value=row['MoM_Pct'])
        ws2.cell(row=row_i, column=4, value=row['YoY_Pct'])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def get_signal(pct):
    if abs(pct) >= 4.0:
        return "🔴", "signal-box-red", "Strong Review Recommended", "badge"
    elif abs(pct) >= 2.0:
        return "🟡", "signal-box-amber", "Selective Review Warranted", "badge-yellow"
    else:
        return "🟢", "signal-box-green", "Stable — Continue Monitoring", "badge-green"


def render_signal_box(pct, sv, ev, start_period, end_period, city, latest_yoy, frequency="Monthly"):
    icon, border_cls, title, badge_cls = get_signal(pct)
    if abs(pct) >= 4.0:
        desc = f"CPI has risen <strong>{pct:+.2f}%</strong> ({start_period}→{end_period}), exceeding the 4% threshold. This supports a formal pricing adjustment proposal."
        badge_label = "≥ 4% Threshold"
    elif abs(pct) >= 2.0:
        desc = f"CPI has moved <strong>{pct:+.2f}%</strong> ({start_period}→{end_period}). Consider selective rate adjustments for CPI-sensitive categories."
        badge_label = "2–4% Range"
    else:
        desc = f"CPI has moved <strong>{pct:+.2f}%</strong> ({start_period}→{end_period}), within acceptable tolerance. No immediate pricing action required."
        badge_label = "< 2% Stable"

    formula_note = f"Formula: (({ev:.2f} − {sv:.2f}) ÷ {sv:.2f}) × 100 = <strong>{pct:+.4f}%</strong> &nbsp;|&nbsp; ABS Official YoY: <strong>{latest_yoy}%</strong>"

    st.markdown(f"""
    <div class="signal-box {border_cls}">
        <div class="signal-icon">{icon}</div>
        <div>
            <div class="signal-title">{title} <span class="badge {badge_cls}">{badge_label}</span></div>
            <div class="signal-desc">{desc}</div>
            <div class="signal-formula">{formula_note}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def render_kpi_row(calc, city_df, city, start_period, end_period, frequency="Monthly", compact=False, group=None):
    pct  = calc['pct_change']
    mv   = calc['movement']
    sv   = calc['start_val']
    ev   = calc['end_val']
    latest_yoy = city_df.iloc[-1]['YoY_Pct']
    latest_mom = city_df.iloc[-1]['MoM_Pct']
    prev_period   = city_df.iloc[-2]['Period']
    n_periods = len(city_df[
        (city_df['Date'] >= city_df[city_df['Period'] == start_period]['Date'].values[0]) &
        (city_df['Date'] <= city_df[city_df['Period'] == end_period]['Date'].values[0])
    ])
    up = pct > 0
    pct_cls = "kpi-red" if up else "kpi-green"
    arrow   = "▲" if up else "▼"
    period_unit = "quarters" if frequency == "Quarterly" else "months"
    # compact class reduces font-size to prevent wrapping in dual-column layout
    c = "kpi kpi-compact" if compact else "kpi"
    wrap_cls = f"kpi-wrap kpi-wrap-{group}" if group else "kpi-wrap"

    st.markdown(f"""
    <div class="{wrap_cls}">
      <div class="{c}">
        <div class="kpi-label">Start CPI</div>
        <div class="kpi-val">{sv:.2f}</div>
        <div class="kpi-sub">{start_period}</div>
      </div>
      <div class="{c}">
        <div class="kpi-label">End CPI</div>
        <div class="kpi-val">{ev:.2f}</div>
        <div class="kpi-sub">{end_period}</div>
      </div>
      <div class="{c} {'kpi-red' if mv > 0 else 'kpi-green'}">
        <div class="kpi-label">Movement</div>
        <div class="kpi-val">{mv:+.2f}</div>
        <div class="kpi-sub">{n_periods} {period_unit}</div>
      </div>
      <div class="{c} {pct_cls}">
        <div class="kpi-label">% Change</div>
        <div class="kpi-val">{arrow} {abs(pct):.1f}%</div>
        <div class="kpi-sub">{start_period}–{end_period}</div>
      </div>
      <div class="{c} kpi-hero {'kpi-red' if (latest_yoy or 0)>0 else 'kpi-green'}">
        <div class="kpi-label">Latest YoY</div>
        <div class="kpi-val">{'▲' if (latest_yoy or 0)>0 else '▼'} {abs(latest_yoy or 0):.1f}%</div>
        <div class="kpi-sub">vs prior year</div>
      </div>
      <div class="{c} {'kpi-red' if (latest_mom or 0)>0 else 'kpi-green'}">
        <div class="kpi-label">{'Latest QoQ' if frequency == 'Quarterly' else 'Latest MoM'}</div>
        <div class="kpi-val">{'▲' if (latest_mom or 0)>0 else '▼'} {abs(latest_mom or 0):.1f}%</div>
        <div class="kpi-sub">vs {prev_period}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)


def filter_city_df(city_df, start_period, end_period):
    start_date = city_df[city_df['Period'] == start_period]['Date'].values[0]
    end_date   = city_df[city_df['Period'] == end_period]['Date'].values[0]
    return city_df[(city_df['Date'] >= start_date) & (city_df['Date'] <= end_date)].copy()


# ─── BRAND LOGO ───────────────────────────────────────────────────────────────

# Set to False if your logo is a clean white-on-transparent mark that should sit
# directly on the red header with no white plate behind it.
LOGO_ON_WHITE_PLATE = True

def _load_logo_html() -> str:
    """
    Return an <img> tag for the Australia Post logo, base64-embedded so it
    travels with the app. Picks up ANY image file dropped into aw-main/assets/
    (any filename, any common format), preferring vector (SVG) then PNG.
    Falls back to the 📮 emoji placeholder when the folder has no image.

    Not cached on purpose — a newly-added file shows on the next rerun/refresh
    without needing a full restart.
    """
    import base64
    from pathlib import Path

    assets = Path(__file__).parent / "assets"
    mime_by_ext = {
        ".svg": "image/svg+xml", ".png": "image/png",
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".webp": "image/webp", ".gif": "image/gif",
    }
    priority = [".svg", ".png", ".webp", ".jpg", ".jpeg", ".gif"]
    css_class = "ap-logo-img" if LOGO_ON_WHITE_PLATE else "ap-logo-img ap-logo-bare"

    if assets.is_dir():
        imgs = [f for f in assets.iterdir()
                if f.is_file() and f.suffix.lower() in mime_by_ext]
        imgs.sort(key=lambda f: priority.index(f.suffix.lower()))
        if imgs:
            f = imgs[0]
            mime = mime_by_ext[f.suffix.lower()]
            b64 = base64.b64encode(f.read_bytes()).decode()
            return (f'<img src="data:{mime};base64,{b64}" alt="Australia Post" '
                    f'class="{css_class}"/>')
    # No image in assets/ — keep the emoji placeholder
    return '<div class="ap-logo">📮</div>'


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    # ── Apply theme (light default) before anything renders ───────────────
    appearance = st.session_state.get("appearance", "☀️ Light")
    apply_theme("Light" if "Light" in appearance else "Dark")

    # ── Load data upfront (cached) ────────────────────────────────────────
    with st.spinner("Fetching latest CPI data from ABS..."):
        df_monthly = fetch_monthly()

    latest_label = df_monthly['Period'].iloc[-1] if df_monthly is not None else "—"

    st.markdown(f"""
    <div class="ap-header">
        {_load_logo_html()}
        <div>
            <h1>CPI Calculator <span style="font-weight:300;opacity:0.75;">| Internal Use Only</span></h1>
            <div class="sub">
                <span>Data &amp; Pricing Executive Tool</span>
                <span style="opacity:0.4;">·</span>
                <span>ABS 6401.0</span>
                <span style="opacity:0.4;">·</span>
                <span>Base Sep-2025 = 100</span>
                <span style="opacity:0.4;">·</span>
                <span>Latest: <strong>{latest_label}</strong></span>
                <span style="opacity:0.4;">·</span>
                <span class="live-dot"></span>
                <span>Live ABS Data</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if df_monthly is None:
        st.error("❌ Failed to load monthly CPI data from ABS. Please check your internet connection.")
        st.stop()

    # ── SIDEBAR ───────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown('<p class="sec-title-sidebar">Series Mode</p>', unsafe_allow_html=True)
        series_mode = st.radio(
            "Series mode",
            ["📅 Monthly only", "📆 Quarterly   ·   ID : A2325846C"],
            index=0,
            label_visibility="collapsed",
            help="Monthly = ABS 640101 (~68% basket) | Quarterly = ABS 6401017 (full basket)"
        )

        use_monthly   = "Monthly"   in series_mode
        use_quarterly = "Quarterly" in series_mode

        # Fetch quarterly data early so periods are available for Period Filter
        df_quarterly = None
        if use_quarterly:
            df_quarterly = fetch_quarterly()

        active_df = df_monthly if use_monthly else (df_quarterly if df_quarterly is not None else df_monthly)
        periods = active_df['Period'].tolist() if active_df is not None else []
        if not periods:
            st.error("No period data available.")
            st.stop()

        # ── Period Filter (directly under Series Mode) ────────────────────
        st.markdown('<p class="sec-title-sidebar">📅 Period Filter</p>', unsafe_allow_html=True)
        snapshot_period = st.selectbox(
            "📅 Select Period", periods,
            index=len(periods)-1,
            label_visibility="collapsed",
            help="View CPI snapshot for this single period across all regions"
        )

        # ── Region ────────────────────────────────────────────────────────
        st.markdown('<p class="sec-title-sidebar">📍 Region</p>', unsafe_allow_html=True)
        city = st.selectbox("📍 Region", CITIES, index=0, label_visibility="collapsed")

        # ── View Mode ─────────────────────────────────────────────────────
        st.markdown('<p class="sec-title-sidebar">View Mode</p>', unsafe_allow_html=True)
        view_mode = st.radio(
            "View mode",
            ["📊 Period Range", "📅 Single Month Snapshot"] if use_monthly
            else ["📊 Period Range", "📅 Single Quarter Snapshot"],
            index=1,
            label_visibility="collapsed",
            help="Period Range: compare two dates | Single Snapshot: one period, all regions"
        )
        single_snapshot = "Snapshot" in view_mode

        # Period Range — show start/end pickers inline below View Mode
        if not single_snapshot:
            col_s, col_e = st.columns(2)
            with col_s:
                start_period = st.selectbox("Start", periods,
                                            index=max(0, len(periods)-13),
                                            help="Base period for comparison")
            with col_e:
                end_period = st.selectbox("End", periods,
                                          index=len(periods)-1,
                                          help="Comparison end period")
        else:
            start_period = snapshot_period
            end_period   = snapshot_period

        # ── Data Sources ──────────────────────────────────────────────────
        st.markdown('<p class="sec-title-sidebar">Data Sources</p>', unsafe_allow_html=True)
        if use_monthly:
            with st.expander("📅 Monthly CPI Indicator", expanded=False):
                st.markdown(f"""<div class="data-status">
                ABS 6401.0 Table 1 · Live · 640101.xlsx<br>
                <strong>{df_monthly['Period'].iloc[0]}</strong> → <strong>{df_monthly['Period'].iloc[-1]}</strong>
                &nbsp;({len(df_monthly)} months)
                </div>""", unsafe_allow_html=True)
        if use_quarterly and df_quarterly is not None:
            with st.expander("📆 Quarterly CPI (Full Basket)", expanded=False):
                st.markdown(f"""<div class="data-status">
                ABS 6401.0 Table 17 · Full basket · 6401017.xlsx<br>
                <strong>{df_quarterly['Period'].iloc[0]}</strong> → <strong>{df_quarterly['Period'].iloc[-1]}</strong>
                &nbsp;({len(df_quarterly)} quarters)
                </div>""", unsafe_allow_html=True)

        # ── Display Options ───────────────────────────────────────────────
        st.markdown('<p class="sec-title-sidebar">Display Options</p>', unsafe_allow_html=True)
        # "Overlay all regions" only applies in Period Range mode
        show_all_regions = (
            st.checkbox("Overlay all regions on index chart", value=False)
            if not single_snapshot else False
        )
        show_yoy_chart = st.checkbox("Show YoY % chart", value=True)
        show_pricing   = st.checkbox("Pricing guidance panel", value=True)
        show_rollup    = (
            st.checkbox("Monthly → Quarterly rollup", value=False,
                        help="Group monthly data by quarter with QoQ summary")
            if (use_monthly and not single_snapshot) else False
        )

        # ── Appearance ────────────────────────────────────────────────────
        st.markdown('<p class="sec-title-sidebar">🎨 Appearance</p>', unsafe_allow_html=True)
        st.radio(
            "Appearance", ["☀️ Light", "🌙 Dark"],
            key="appearance", horizontal=True,
            label_visibility="collapsed",
            help="Switch between light and dark themes",
        )


    # ── VALIDATION ────────────────────────────────────────────────────────
    if not single_snapshot:
        s_idx = periods.index(start_period)
        e_idx = periods.index(end_period)
        if s_idx >= e_idx:
            st.error("⚠️ Start period must be before End period.")
            st.stop()

    # ══════════════════════════════════════════════════════════════════════
    # SINGLE SERIES MODE
    # ══════════════════════════════════════════════════════════════════════
    df = df_monthly if use_monthly else df_quarterly
    frequency = df['Frequency'].iloc[0] if df is not None else "Monthly"

    if single_snapshot:
        _render_snapshot_month(df, snapshot_period, frequency, selected_city=city,
                               show_yoy_chart=show_yoy_chart, show_pricing=show_pricing)
    else:
        _render_single(df, city, start_period, end_period,
                       show_all_regions, show_yoy_chart, show_pricing, frequency,
                       show_rollup=show_rollup)


# ─── SINGLE PERIOD SNAPSHOT RENDERER ────────────────────────────────────────

def _render_snapshot_month(df, period, frequency="Monthly", selected_city="Australia",
                           show_yoy_chart=True, show_pricing=True):
    """Show all 9 regions side-by-side for one selected period."""
    row = df[df['Period'] == period]
    if row.empty:
        st.error(f"No data for period {period}.")
        return
    row = row.iloc[0]

    freq_pill = (
        '<span class="series-pill pill-quarterly">Quarterly</span>'
        if frequency == "Quarterly"
        else '<span class="series-pill pill-monthly">Monthly</span>'
    )
    pop_label = "QoQ" if frequency == "Quarterly" else "MoM"
    series_color = QUARTERLY_COLOR if frequency == "Quarterly" else MONTHLY_COLOR

    # YoY colour follows the pricing-signal thresholds
    city_yoy = row.get(f'{selected_city}_YoY')
    yoy_signal_color = (MONTHLY_COLOR if (city_yoy or 0) >= 4
                        else ("#D97706" if (city_yoy or 0) >= 2 else "#16A34A"))

    deriv_html = ""
    if show_pricing:
        deriv_html = ('<div style="font-size:0.70rem;color:var(--ink-600);margin-top:5px;'
                      'font-weight:500;">Derivation in Calculation Evidence below ↓</div>')

    yoy_disp = f"{city_yoy:+.1f}%" if city_yoy is not None else "—"
    st.markdown(f"""
    <div class="dual-header">
        <div style="width:10px;height:10px;background:{series_color};border-radius:50%;flex-shrink:0;margin-top:4px;"></div>
        <div class="dual-header-item">
            <div class="dual-header-label">Period {freq_pill}</div>
            <div class="dual-header-val" style="color:{series_color};">{period}</div>
        </div>
        <div style="width:1px;background:var(--border);align-self:stretch;margin:0 4px;flex-shrink:0;"></div>
        <div class="dual-header-item">
            <div class="dual-header-label">Region</div>
            <div class="dual-header-val" style="color:var(--ink-700);">{selected_city}</div>
        </div>
        <div style="width:1px;background:var(--border);align-self:stretch;margin:0 4px;flex-shrink:0;"></div>
        <div class="dual-header-item">
            <div class="dual-header-label">YoY %</div>
            <div class="dual-header-val" style="color:{yoy_signal_color};">{yoy_disp}</div>
            {deriv_html}
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Build region comparison table
    snap_rows = []
    for city in CITIES:
        idx = row.get(f'{city}_Index')
        yoy = row.get(f'{city}_YoY')
        pop = row.get(f'{city}_MoM')
        signal = "🔴" if (yoy or 0) >= 4 else ("🟡" if (yoy or 0) >= 2 else "🟢")
        snap_rows.append({
            'Region':           city,
            'CPI Index':        round(float(idx), 2) if idx is not None else None,
            f'YoY %':           round(float(yoy), 2) if yoy is not None else None,
            f'{pop_label} %':   round(float(pop), 2) if pop is not None else None,
            'Signal':           signal,
        })

    snap_df = pd.DataFrame(snap_rows)

    # Evidence panel for selected city
    _render_yoy_evidence(df, selected_city, period, frequency, default_expanded=True)

    # ── Fix 1 & 2: Sort both charts by YoY descending ─────────────────────
    sorted_snap = snap_df.sort_values('YoY %', ascending=False).reset_index(drop=True)
    sorted_cities = sorted_snap['Region'].tolist()

    # ── Australia national average for delta labels ─────────────────────────
    aus_val_series = snap_df.loc[snap_df['Region'] == 'Australia', 'CPI Index']
    aus_val = float(aus_val_series.values[0]) if not aus_val_series.empty else None

    # ── CPI Index chart (sorted) ──────────────────────────────────────────
    st.markdown(f'<p class="sec-title">CPI Index by Region — {period}</p>', unsafe_allow_html=True)

    idx_bar_colors = [series_color if c == selected_city else '#E8A0A8' for c in sorted_cities]
    idx_line_colors = ['white' if c == selected_city else 'rgba(0,0,0,0)' for c in sorted_cities]
    idx_line_widths = [2.5 if c == selected_city else 0 for c in sorted_cities]

    # Fix 3: "vs Aus" delta as second line of bar label
    if aus_val:
        idx_labels = [
            f"{v:.2f}\n{(v - aus_val):+.2f} vs Aus" if c != 'Australia' else f"{v:.2f}\n(National Avg)"
            for c, v in zip(sorted_cities, sorted_snap['CPI Index'])
        ]
    else:
        idx_labels = [f"{v:.2f}" for v in sorted_snap['CPI Index']]

    fig = go.Figure(go.Bar(
        x=sorted_snap['Region'],
        y=sorted_snap['CPI Index'],
        marker=dict(
            color=idx_bar_colors,
            line=dict(color=idx_line_colors, width=idx_line_widths),
        ),
        text=idx_labels,
        textposition="outside",
        textfont=dict(size=9, color=CHART_FONT, weight="bold"),
        hovertemplate="<b>%{x}</b><br>Index: %{y:.2f}<extra></extra>",
    ))
    fig.add_hline(y=100, line_dash="dash", line_color="#ccc", line_width=1,
                  annotation=dict(text="Base Sep-2025 = 100", font=dict(size=8, color=CHART_FONT, weight="bold"),
                                  bgcolor=CHART_PILL, bordercolor="#E5E7EB", borderwidth=1, borderpad=2),
                  annotation_position="bottom left")
    if aus_val and selected_city != 'Australia':
        fig.add_hline(y=aus_val, line_dash="dot", line_color=series_color, line_width=1,
                      annotation=dict(text=f"Aus avg {aus_val:.2f}", font=dict(size=8, color=series_color),
                                      bgcolor=CHART_PILL, bordercolor=series_color, borderwidth=1, borderpad=2),
                      annotation_position="bottom right")

    # Fix 6: "◀ Selected" annotation above selected city bar in index chart
    sel_idx_pos = sorted_cities.index(selected_city) if selected_city in sorted_cities else None
    if sel_idx_pos is not None:
        sel_y = float(sorted_snap.loc[sorted_snap['Region'] == selected_city, 'CPI Index'].values[0])
        fig.add_annotation(x=selected_city, y=sel_y,
                           text="▼ Selected",
                           showarrow=False, yanchor="bottom",
                           yshift=32, font=dict(size=8, color=series_color), bgcolor=CHART_PILL,
                           bordercolor=series_color, borderwidth=1, borderpad=3)

    idx_min = sorted_snap['CPI Index'].min() if sorted_snap['CPI Index'].notna().any() else 95
    idx_max = sorted_snap['CPI Index'].max() if sorted_snap['CPI Index'].notna().any() else 105
    fig.update_layout(
        xaxis=dict(gridcolor=CHART_GRID, tickangle=-30, categoryorder="array",
                   categoryarray=sorted_cities),
        yaxis=dict(title="Index", gridcolor=CHART_GRID,
                   range=[idx_min * 0.992, idx_max * 1.012]),
        plot_bgcolor=CHART_BG, paper_bgcolor=CHART_BG,
        height=340, showlegend=False,
        margin=dict(l=40, r=120, t=30, b=55),
        font=dict(family="IBM Plex Sans", color=CHART_FONT, weight="bold"),
    )
    themed_plotly(fig, use_container_width=True, config={"displayModeBar": False})

    # ── YoY chart (same sorted order) ────────────────────────────────────
    if not show_yoy_chart:
        return
    st.markdown(f'<p class="sec-title">YoY % Change by Region — {period}</p>', unsafe_allow_html=True)

    yoy_colors = [
        MONTHLY_COLOR if (v or 0) >= 4 else ("#F59E0B" if (v or 0) >= 2 else "#16A34A")
        for v in sorted_snap['YoY %']
    ]
    yoy_line_colors = ['white' if c == selected_city else 'rgba(0,0,0,0)' for c in sorted_cities]
    yoy_line_widths = [2.5 if c == selected_city else 0 for c in sorted_cities]
    yoy_max = sorted_snap['YoY %'].max() if sorted_snap['YoY %'].notna().any() else 5
    y_top_snap = max(yoy_max * 1.28, 6.0)

    fig2 = go.Figure()
    # Main bars
    fig2.add_trace(go.Bar(
        x=sorted_snap['Region'], y=sorted_snap['YoY %'],
        showlegend=False,
        marker=dict(color=yoy_colors, line=dict(color=yoy_line_colors, width=yoy_line_widths)),
        text=sorted_snap['YoY %'].apply(lambda v: f"{v:+.1f}%" if v is not None else "—"),
        textposition="outside", textfont=dict(size=10, color=CHART_FONT, weight="bold"),
        hovertemplate="<b>%{x}</b><br>YoY: %{y:+.1f}%<extra></extra>",
    ))
    # Fix 4: colour legend — 3 dummy traces
    for lbl, clr in [("≥4% — Strong Review", MONTHLY_COLOR),
                     ("2–4% — Monitor", "#F59E0B"),
                     ("<2% — Stable", "#16A34A")]:
        fig2.add_trace(go.Bar(x=[], y=[], name=lbl, marker_color=clr, showlegend=True))

    # Threshold reference lines — labels as white pill chips to stay legible over bars
    fig2.add_hline(y=4.0, line_dash="dash", line_color=MONTHLY_COLOR, line_width=1.5,
                   annotation=dict(text="4% Review threshold", font=dict(color=MONTHLY_COLOR, size=9),
                                   bgcolor=CHART_PILL, bordercolor=MONTHLY_COLOR, borderwidth=1, borderpad=3),
                   annotation_position="top left")
    fig2.add_hline(y=2.0, line_dash="dot", line_color="#F59E0B", line_width=1,
                   annotation=dict(text="2% Monitor level", font=dict(color="#FBBF24", size=9),
                                   bgcolor=CHART_PILL, bordercolor="#F59E0B", borderwidth=1, borderpad=3),
                   annotation_position="bottom left")

    # Fix 6: "◀ Selected" annotation above selected city bar in YoY chart
    if selected_city in sorted_cities:
        sel_yoy = sorted_snap.loc[sorted_snap['Region'] == selected_city, 'YoY %'].values
        if len(sel_yoy) and sel_yoy[0] is not None:
            fig2.add_annotation(x=selected_city, y=float(sel_yoy[0]),
                                text="▼ Selected",
                                showarrow=False, yanchor="bottom",
                                yshift=22, font=dict(size=8, color=series_color),
                                bgcolor=CHART_PILL, bordercolor=series_color,
                                borderwidth=1, borderpad=3)

    fig2.update_layout(
        xaxis=dict(gridcolor=CHART_GRID, tickangle=-30, categoryorder="array",
                   categoryarray=sorted_cities),
        yaxis=dict(title="YoY %", gridcolor=CHART_GRID, range=[0, y_top_snap]),
        plot_bgcolor=CHART_BG, paper_bgcolor=CHART_BG,
        height=380, barmode="group",
        legend=dict(orientation="h", y=1.08, x=0, font=dict(size=9)),
        margin=dict(l=40, r=40, t=55, b=55),
        font=dict(family="IBM Plex Sans", color=CHART_FONT, weight="bold"),
        showlegend=True,
    )
    themed_plotly(fig2, use_container_width=True, config={"displayModeBar": False})

    # Region table
    st.markdown('<p class="sec-title">All Regions — Detail</p>', unsafe_allow_html=True)

    def color_pct(val):
        if val is None or not isinstance(val, (int, float)):
            return ''
        return f'color: {POS_COLOR}; font-weight:600' if val > 0 else (f'color: {NEG_COLOR}; font-weight:600' if val < 0 else '')

    styled = (snap_df.style
        .format({'CPI Index': '{:.2f}',
                 'YoY %':         lambda v: f'{v:+.2f}%' if v is not None else '—',
                 f'{pop_label} %': lambda v: f'{v:+.2f}%' if v is not None else '—'})
        .map(color_pct, subset=['YoY %', f'{pop_label} %'])
    )
    render_themed_table(styled)

    # Data note
    aus_idx = row.get('Australia_Index')
    aus_yoy = row.get('Australia_YoY')
    st.markdown(f"""
    <div class="abs-note">
        <strong>Australia (national)</strong> — CPI Index: <strong>{aus_idx:.2f}</strong>
        &nbsp;|&nbsp; YoY: <strong>{f"{aus_yoy:+.1f}%" if aus_yoy is not None else "—"}</strong>
        &nbsp;|&nbsp; Period: <strong>{period}</strong>
        &nbsp;|&nbsp; Source: ABS 6401.0 (640101) · Base Sep-2025 = 100
    </div>
    """, unsafe_allow_html=True)

    # ── Quarter context (monthly snapshot only) ───────────────────────────
    if frequency == "Monthly":
        sel_date  = df[df['Period'] == period]['Date'].values[0]
        sel_ts    = pd.Timestamp(sel_date)
        q_num     = (sel_ts.month - 1) // 3 + 1
        quarter   = f"Q{q_num}-{sel_ts.year}"
        q_months  = {1:[1,2,3], 2:[4,5,6], 3:[7,8,9], 4:[10,11,12]}[q_num]
        q_end_month = q_months[-1]

        # All months in this quarter that exist in our data
        months_in_q = df[
            (df['Date'].dt.year  == sel_ts.year) &
            (df['Date'].dt.month.isin(q_months))
        ][['Period','Date']].copy()

        # Quarterly CPI for this quarter (if quarter-end month exists)
        df_q      = derive_quarterly_from_monthly(df)
        q_row     = df_q[df_q['Period'] == quarter]
        q_aus_idx = q_row['Australia_Index'].values[0] if not q_row.empty else None
        q_aus_qoq = q_row['Australia_MoM'].values[0]  if not q_row.empty else None
        q_aus_yoy = q_row['Australia_YoY'].values[0]  if not q_row.empty else None

        month_names = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
                       7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
        expected_labels = [f"{month_names[m]}-{sel_ts.year}" for m in q_months]
        available = set(months_in_q['Period'].tolist())

        dots = ""
        for lbl in expected_labels:
            if lbl == period:
                dots += f'<span style="display:inline-block;background:#C8102E;color:white;border-radius:6px;padding:4px 10px;margin:3px;font-size:0.78rem;font-weight:700;">{lbl} ◀ selected</span>'
            elif lbl in available:
                is_qend = lbl.startswith(month_names[q_end_month])
                bg = "#1D4ED8" if is_qend else DOT_BG
                fg = "white" if is_qend else DOT_FG
                suffix = " ◇ Q-End" if is_qend else ""
                dots += f'<span style="display:inline-block;background:{bg};color:{fg};border-radius:6px;padding:4px 10px;margin:3px;font-size:0.78rem;">{lbl}{suffix}</span>'
            else:
                dots += f'<span style="display:inline-block;background:{DOT_ND_BG};color:{DOT_ND_FG};border:1px dashed {DOT_ND_BORDER};border-radius:6px;padding:4px 10px;margin:3px;font-size:0.78rem;">{lbl} (no data)</span>'

        q_idx_str = f"{q_aus_idx:.2f}" if q_aus_idx is not None else "—"
        q_qoq_str = f"{q_aus_qoq:+.2f}%" if q_aus_qoq is not None else "pending"
        q_yoy_str = f"{q_aus_yoy:+.1f}%" if q_aus_yoy is not None else "—"
        q_end_avail = q_end_month in months_in_q['Date'].dt.month.values if not months_in_q.empty else False
        q_status = "✅ Complete" if q_end_avail else "⏳ In progress"

        st.markdown(f"""
        <div style="background:rgba(96,165,250,0.10);border:1px solid rgba(96,165,250,0.28);border-radius:var(--radius);box-shadow:var(--shadow-sm);padding:var(--s-4) var(--s-5);margin-top:var(--s-4);">
            <div style="font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:var(--blue);margin-bottom:var(--s-3);">
                Quarter Context — {quarter} &nbsp; <span style="font-weight:400;color:var(--ink-600);">{q_status}</span>
            </div>
            <div style="margin-bottom:var(--s-3);">{dots}</div>
            <div style="font-size:0.82rem;color:var(--ink-700);display:flex;gap:var(--s-5);flex-wrap:wrap;">
                <span><strong style="color:var(--blue);">Quarter:</strong> {quarter}</span>
                <span><strong style="color:var(--blue);">Q-End Index (Australia):</strong> {q_idx_str}</span>
                <span><strong style="color:var(--blue);">QoQ:</strong> {q_qoq_str}</span>
                <span><strong style="color:var(--blue);">YoY at Q-End:</strong> {q_yoy_str}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)


# ─── CALCULATION EVIDENCE ────────────────────────────────────────────────────

def _render_yoy_evidence(df, city, period, frequency, default_expanded=True):
    """
    Step-by-step YoY and MoM/QoQ calculation evidence for a given city and period.
    Shows exactly how ABS index numbers translate to the published percentage figures.
    """
    idx_col = f'{city}_Index'
    yoy_col = f'{city}_YoY'
    mom_col = f'{city}_MoM'

    row = df[df['Period'] == period]
    if row.empty or idx_col not in df.columns:
        return
    row = row.iloc[0]

    curr_idx = row.get(idx_col)
    curr_yoy = row.get(yoy_col)
    if curr_idx is None:
        return
    curr_f = float(curr_idx)

    # ── Prior-year period ────────────────────────────────────────────────
    if frequency == 'Quarterly':
        q, yr     = period[1], int(period[3:])
        prior_yr  = f"Q{q}-{yr - 1}"
        mom_lbl   = "QoQ"
        freq_word = "Quarterly"
        source    = "ABS 6401.0 Table 17 — CPI All Groups (100% full basket)"
    else:
        parts     = period.split('-')
        prior_yr  = f"{parts[0]}-{int(parts[1]) - 1}"
        mom_lbl   = "MoM"
        freq_word = "Monthly"
        source    = "ABS 6401.0 Table 1 — Monthly CPI Indicator (~68% basket)"

    prior_yr_row = df[df['Period'] == prior_yr]
    prior_yr_idx = (float(prior_yr_row[idx_col].values[0])
                    if not prior_yr_row.empty and prior_yr_row[idx_col].values[0] is not None
                    else None)

    # ── YoY block ────────────────────────────────────────────────────────
    if prior_yr_idx:
        diff   = curr_f - prior_yr_idx
        calc   = diff / prior_yr_idx * 100
        rounds = (f'<br><span style="font-size:0.78rem;color:var(--ink-600);font-family:inherit;">→ ABS official: </span>'
                  f'<strong style="font-size:1.15rem;color:#FF5A73;">{curr_yoy:.1f}%</strong>') if curr_yoy is not None else ""
        yoy_block = f"""
        <div style="display:flex;gap:20px;align-items:flex-start;flex-wrap:wrap;">
          <div style="flex:0 0 auto;">
            <div class="ev-row">
              <span class="ev-step">①</span>
              <span class="ev-lbl">{period} &nbsp;<em style="color:#9CA3AF;">(current)</em></span>
              <span class="ev-val">{curr_f:.2f}</span>
            </div>
            <div class="ev-row">
              <span class="ev-step">②</span>
              <span class="ev-lbl">{prior_yr} &nbsp;<em style="color:#9CA3AF;">(one year ago)</em></span>
              <span class="ev-val">{prior_yr_idx:.2f}</span>
            </div>
          </div>
          <div class="ev-formula" style="flex:1;min-width:260px;margin-top:0;">
            ③ &nbsp;({curr_f:.2f} − {prior_yr_idx:.2f}) ÷ {prior_yr_idx:.2f} × 100<br>
            &nbsp;&nbsp;&nbsp;&nbsp;= {diff:.4f} ÷ {prior_yr_idx:.2f} × 100<br>
            &nbsp;&nbsp;&nbsp;&nbsp;= <strong style="color:#FF5A73;font-size:1.05rem;">{calc:+.2f}%</strong>{rounds}
          </div>
        </div>"""
    else:
        yoy_block = f"<p style='color:#9CA3AF;font-size:0.8rem;'>Prior year data ({prior_yr}) not yet available.</p>"

    # ── Recent trend table (last 6 periods, selected row highlighted) ────
    recent = df[['Period', idx_col, yoy_col, mom_col]].tail(6).copy()
    recent = recent.rename(columns={idx_col: 'Index', yoy_col: 'YoY %', mom_col: f'{mom_lbl} %'})

    def _hl(r):
        base = 'background-color:rgba(255,90,115,0.16);font-weight:700;' if r['Period'] == period else ''
        return [base] * len(r)

    styled = (recent.style
        .format({'Index':    lambda v: f'{float(v):.2f}' if v is not None else '—',
                 'YoY %':    lambda v: f'{float(v):+.1f}%' if v is not None else '—',
                 f'{mom_lbl} %': lambda v: f'{float(v):+.1f}%' if v is not None else '—'})
        .apply(_hl, axis=1))

    spark_color = QUARTERLY_COLOR if frequency == 'Quarterly' else MONTHLY_COLOR

    with st.expander(f"📐  {city}  ·  {period}  —  Calculation Evidence", expanded=default_expanded):
        st.markdown(f"""
        <div>
          <div class="ev-heading">Year-on-Year ({freq_word} vs Same Period Prior Year)</div>
          {yoy_block}
        </div>
        <div class="ev-source">📎 {source} &nbsp;·&nbsp; Base: Sep-2025 = 100 &nbsp;·&nbsp; All calculations use ABS published index numbers</div>
        """, unsafe_allow_html=True)

        st.markdown(f'<p class="ev-heading" style="margin-top:16px;">Recent {city} Index History — last 6 periods</p>', unsafe_allow_html=True)

        spark_x = recent['Period'].tolist()
        spark_y = [float(v) if v is not None else None for v in recent['Index']]
        dot_colors = [spark_color if p == period else '#E8A0A8' for p in spark_x]
        dot_sizes  = [10 if p == period else 6 for p in spark_x]
        fig_spark = go.Figure(go.Scatter(
            x=spark_x, y=spark_y,
            mode='lines+markers',
            line=dict(color=spark_color, width=2.5),
            marker=dict(color=dot_colors, size=dot_sizes, line=dict(color='white', width=1.5)),
            hovertemplate="<b>%{x}</b><br>Index: %{y:.2f}<extra></extra>",
        ))
        fig_spark.update_layout(
            height=175, margin=dict(l=40, r=20, t=10, b=30),
            plot_bgcolor=CHART_BG, paper_bgcolor=CHART_BG,
            xaxis=dict(type='category', tickfont=dict(size=11, color=CHART_FONT, weight="bold"), gridcolor=CHART_GRID, fixedrange=True),
            yaxis=dict(tickfont=dict(size=11, color=CHART_FONT, weight="bold"), gridcolor=CHART_GRID, fixedrange=True),
            showlegend=False, font=dict(family='IBM Plex Sans', color=CHART_FONT, weight="bold"),
        )
        themed_plotly(fig_spark, use_container_width=True, config={"displayModeBar": False})
        render_themed_table(styled)


# ─── SINGLE SERIES RENDERER ──────────────────────────────────────────────────

def _render_single(df, city, start_period, end_period,
                   show_all_regions, show_yoy_chart, show_pricing, frequency="Monthly",
                   show_rollup=False):
    calc     = calc_custom_change(df, city, start_period, end_period)
    city_df  = get_city_df(df, city)
    pct, sv, ev = calc['pct_change'], calc['start_val'], calc['end_val']
    latest_yoy = city_df.iloc[-1]['YoY_Pct']
    series_color = QUARTERLY_COLOR if frequency == "Quarterly" else MONTHLY_COLOR

    render_kpi_row(calc, city_df, city, start_period, end_period, frequency)

    if show_pricing:
        st.markdown('<p class="sec-title">Pricing Signal</p>', unsafe_allow_html=True)
        render_signal_box(pct, sv, ev, start_period, end_period, city, latest_yoy, frequency)

    # Evidence panel — shows how the Latest YoY figure is derived from ABS index numbers
    latest_period = city_df.iloc[-1]['Period']
    _render_yoy_evidence(df, city, latest_period, frequency)

    sub = filter_city_df(city_df, start_period, end_period)

    chart_col1, chart_col2 = (st.columns(2) if show_yoy_chart else (st.container(), None))

    with (chart_col1 if show_yoy_chart else st):
        st.markdown(f'<p class="sec-title">CPI Index Trend — {city}</p>', unsafe_allow_html=True)
        fig = go.Figure()
        if show_all_regions:
            colors_map = {
                'Australia': MONTHLY_COLOR, 'Sydney': '#60A5FA', 'Melbourne': '#34D399',
                'Brisbane': '#FBBF24', 'Adelaide': '#A78BFA', 'Perth': '#22D3EE',
                'Hobart': '#FB923C', 'Darwin': '#F472B6', 'Canberra': '#94A3B8'
            }
            for c in CITIES:
                c_sub = filter_city_df(get_city_df(df, c), start_period, end_period)
                fig.add_trace(go.Scatter(
                    x=c_sub['Period'], y=c_sub['CPI_Index'], name=c,
                    line=dict(width=2.5 if c == city else 1.5, color=colors_map[c],
                              dash='solid' if c == city else 'dot'),
                    mode="lines",
                ))
        else:
            # Subtle fill between line and baseline (100), not tozeroy
            r, g, b_ch = int(series_color[1:3],16), int(series_color[3:5],16), int(series_color[5:7],16)
            fig.add_trace(go.Scatter(
                x=sub['Period'], y=sub['CPI_Index'], name=city,
                line=dict(color=series_color, width=2.5),
                mode="lines+markers",
                marker=dict(size=7, color=series_color),
                fill="tozeroy",
                fillcolor=f"rgba({r},{g},{b_ch},0.05)",
                hovertemplate="<b>%{x}</b><br>Index: %{y:.2f}<extra></extra>",
            ))
            # Consistent annotation style — same shape, differentiated by colour
            fig.add_annotation(x=start_period, y=sv,
                text=f"<b>{sv:.2f}</b><br>{start_period}",
                showarrow=True, arrowhead=2, ax=-45, ay=-40,
                bgcolor=CHART_PILL, bordercolor="#4A4756", borderwidth=1,
                font=dict(size=10, color=CHART_FONT, weight="bold"))
            fig.add_annotation(x=end_period, y=ev,
                text=f"<b>{ev:.2f}</b><br>{end_period}",
                showarrow=True, arrowhead=2, ax=45, ay=-40,
                bgcolor=CHART_PILL, bordercolor=series_color, borderwidth=1,
                font=dict(size=10, color=series_color))

        # Zoom Y-axis to actual data range with padding (not from 0)
        y_vals = sub['CPI_Index'].dropna()
        y_pad  = max((y_vals.max() - y_vals.min()) * 0.25, 1.5)
        y_min  = round(y_vals.min() - y_pad, 1)
        y_max  = round(y_vals.max() + y_pad, 1)

        fig.add_hline(y=100, line_dash="dash", line_color="#bbb", line_width=1,
                      annotation=dict(text="Base: Sep-2025 = 100",
                                      font=dict(size=9, color=CHART_FONT, weight="bold"),
                                      bgcolor=CHART_PILL, bordercolor="#E5E7EB",
                                      borderwidth=1, borderpad=3),
                      annotation_position="top left")
        fig.update_layout(
            xaxis=dict(tickangle=-45, gridcolor=CHART_GRID, nticks=9),
            yaxis=dict(title="Index", gridcolor=CHART_GRID, range=[y_min, y_max]),
            plot_bgcolor=CHART_BG, paper_bgcolor=CHART_BG,
            hovermode="x unified", height=320,
            legend=dict(orientation="h", y=1.05, x=0),
            margin=dict(l=40, r=90, t=24, b=60),
            font=dict(family="IBM Plex Sans", color=CHART_FONT, weight="bold"),
        )
        themed_plotly(fig, use_container_width=True, config={"displayModeBar": False})

    if show_yoy_chart and chart_col2:
        with chart_col2:
            st.markdown(f'<p class="sec-title">YoY % Change — {city}</p>', unsafe_allow_html=True)
            yoy_sub = sub.dropna(subset=['YoY_Pct'])
            if yoy_sub.empty:
                st.info("YoY data not available for this period range.")
            else:
                fig2 = go.Figure()
                bar_colors = [MONTHLY_COLOR if v >= 4 else ("#F59E0B" if v >= 2 else "#16A34A")
                              for v in yoy_sub['YoY_Pct']]
                fig2.add_trace(go.Bar(
                    x=yoy_sub['Period'], y=yoy_sub['YoY_Pct'],
                    marker_color=bar_colors, showlegend=False,
                    text=yoy_sub['YoY_Pct'].apply(lambda x: f"{x:.1f}%"),
                    textposition="outside", textfont=dict(size=10, color=CHART_FONT, weight="bold"),
                    hovertemplate="<b>%{x}</b><br>YoY: %{y:.1f}%<extra></extra>",
                ))
                # Colour-band legend only — main bars excluded from legend
                for label, clr in [("≥4% Review", MONTHLY_COLOR),
                                    ("2–4% Monitor", "#F59E0B"),
                                    ("<2% Stable", "#16A34A")]:
                    fig2.add_trace(go.Bar(x=[], y=[], name=label,
                                          marker_color=clr, showlegend=True))
                fig2.add_hline(y=4.0, line_dash="dash", line_color=MONTHLY_COLOR, line_width=1.5,
                               annotation=dict(text="4% Review", font=dict(color=MONTHLY_COLOR, size=9),
                                               bgcolor=CHART_PILL, bordercolor=MONTHLY_COLOR,
                                               borderwidth=1, borderpad=3),
                               annotation_position="top left")
                fig2.add_hline(y=2.0, line_dash="dot", line_color="#F59E0B", line_width=1,
                               annotation=dict(text="2% Monitor", font=dict(color="#FBBF24", size=9),
                                               bgcolor=CHART_PILL, bordercolor="#F59E0B",
                                               borderwidth=1, borderpad=3),
                               annotation_position="bottom left")
                y_top = max(yoy_sub['YoY_Pct'].max() * 1.28, 5.5)
                fig2.update_layout(
                    xaxis=dict(tickangle=-45, gridcolor=CHART_GRID, type="category", dtick=2),
                    yaxis=dict(title="%", gridcolor=CHART_GRID, range=[0, y_top]),
                    plot_bgcolor=CHART_BG, paper_bgcolor=CHART_BG,
                    hovermode="x unified", height=320,
                    legend=dict(orientation="h", y=1.05, x=0, font=dict(size=9)),
                    margin=dict(l=40, r=40, t=40, b=60),
                    font=dict(family="IBM Plex Sans", color=CHART_FONT, weight="bold"),
                    showlegend=True,
                )
                themed_plotly(fig2, use_container_width=True, config={"displayModeBar": False})

    _render_table(sub, city_df, start_period, end_period, sv, frequency)
    if show_rollup and frequency == "Monthly":
        _render_monthly_to_quarterly(df, city, start_period, end_period)
    _render_snapshot(df, city)
    _render_exports(city_df, city, start_period, end_period, calc, frequency)
    _render_footer(df, frequency)


# ─── MONTHLY → QUARTERLY ROLLUP ──────────────────────────────────────────────

def _render_monthly_to_quarterly(df_monthly, city, start_period, end_period):
    """
    Group monthly data by quarter and overlay quarterly end-points.
    Shows: combined bar+marker chart, plus a grouped rollup table.
    """
    st.markdown('<p class="sec-title">Monthly → Quarterly Rollup</p>', unsafe_allow_html=True)

    city_m = get_city_df(df_monthly, city)
    sub_m  = filter_city_df(city_m, start_period, end_period).copy()

    # Add quarter label to each monthly row
    sub_m['Quarter'] = sub_m['Date'].apply(
        lambda dt: f"Q{(dt.month - 1) // 3 + 1}-{dt.year}"
    )

    # Derive quarterly series and filter to same date range
    df_q    = derive_quarterly_from_monthly(df_monthly)
    city_q  = get_city_df(df_q, city)
    sub_q   = city_q[
        (city_q['Date'] >= sub_m['Date'].min()) &
        (city_q['Date'] <= sub_m['Date'].max())
    ].copy()

    # ── Combined chart ────────────────────────────────────────────────────
    fig = go.Figure()

    # Monthly bars
    bar_colors = [
        MONTHLY_COLOR if row['Quarter'] == sub_m['Quarter'].iloc[-1] else '#FECDD3'
        for _, row in sub_m.iterrows()
    ]
    fig.add_trace(go.Bar(
        x=sub_m['Period'], y=sub_m['CPI_Index'],
        name="Monthly CPI",
        marker_color=bar_colors,
        opacity=0.75,
        hovertemplate="<b>%{x}</b><br>Index: %{y:.2f}<extra></extra>",
    ))

    # Quarterly end-point markers
    if not sub_q.empty:
        fig.add_trace(go.Scatter(
            x=sub_q['Period'], y=sub_q['CPI_Index'],
            name="Quarter End",
            mode="markers+lines",
            marker=dict(size=14, color=QUARTERLY_COLOR, symbol="diamond",
                        line=dict(color="white", width=2)),
            line=dict(color=QUARTERLY_COLOR, width=2, dash="dash"),
            hovertemplate="<b>%{x}</b><br>Quarter End Index: %{y:.2f}<extra></extra>",
        ))

    fig.add_hline(y=100, line_dash="dash", line_color="#ccc", line_width=1,
                  annotation=dict(text="Base: Sep-2025 = 100", font=dict(size=9, color=CHART_FONT, weight="bold"),
                                  bgcolor=CHART_PILL, bordercolor=CHART_GRID, borderwidth=1, borderpad=2),
                  annotation_position="right")
    fig.update_layout(
        title=dict(
            text=f"{city} — Monthly CPI (bars) with Quarter-End markers (◇)",
            font=dict(size=13, color=CHART_FONT, weight="bold"), x=0
        ),
        xaxis=dict(tickangle=-45, gridcolor=CHART_GRID),
        yaxis=dict(title="CPI Index", gridcolor=CHART_GRID),
        plot_bgcolor=CHART_BG, paper_bgcolor=CHART_BG,
        hovermode="x unified", height=360,
        legend=dict(orientation="h", y=1.08, x=0),
        margin=dict(l=40, r=10, t=50, b=60),
        font=dict(family="IBM Plex Sans", color=CHART_FONT, weight="bold"),
        barmode="group",
    )
    themed_plotly(fig, use_container_width=True)

    # ── Grouped rollup table ──────────────────────────────────────────────
    st.markdown('<p class="sec-title">Quarter Rollup Table</p>', unsafe_allow_html=True)

    q_summary = {}
    for _, qrow in sub_q.iterrows():
        q_summary[qrow['Period']] = {
            'end_index': qrow['CPI_Index'],
            'yoy':       qrow['YoY_Pct'],
            'qoq':       qrow['MoM_Pct'],
        }

    rollup_rows = []
    for quarter, grp in sub_m.groupby('Quarter', sort=False):
        for _, mrow in grp.iterrows():
            rollup_rows.append({
                'Quarter':   quarter,
                'Month':     mrow['Period'],
                'CPI Index': round(mrow['CPI_Index'], 2) if mrow['CPI_Index'] else None,
                'MoM %':     round(mrow['MoM_Pct'], 2)  if mrow['MoM_Pct']  else None,
                'YoY %':     round(mrow['YoY_Pct'], 2)  if mrow['YoY_Pct']  else None,
                'Type':      'month',
            })
        # Quarter summary row
        if quarter in q_summary:
            qs = q_summary[quarter]
            rollup_rows.append({
                'Quarter':   quarter,
                'Month':     f"▶ {quarter} END",
                'CPI Index': round(qs['end_index'], 2) if qs['end_index'] else None,
                'MoM %':     None,
                'YoY %':     round(qs['yoy'], 2) if qs['yoy'] else None,
                'Type':      'quarter',
            })

    rollup_df = pd.DataFrame(rollup_rows)
    display_df = rollup_df[['Quarter', 'Month', 'CPI Index', 'MoM %', 'YoY %']].copy()

    def style_row(row):
        if rollup_df.loc[row.name, 'Type'] == 'quarter':
            return [f'background-color:#DBEAFE; font-weight:700; color:#1D4ED8'] * len(row)
        return [''] * len(row)

    def color_pct(val):
        if val is None or not isinstance(val, (int, float)):
            return ''
        return f'color:{POS_COLOR};font-weight:600' if val > 0 else (f'color:{NEG_COLOR};font-weight:600' if val < 0 else '')

    styled = (display_df.style
        .apply(style_row, axis=1)
        .format({
            'CPI Index': lambda v: f'{v:.2f}' if v is not None else '—',
            'MoM %':     lambda v: f'{v:+.2f}%' if v is not None else '—',
            'YoY %':     lambda v: f'{v:+.2f}%' if v is not None else '—',
        })
        .map(color_pct, subset=['MoM %', 'YoY %'])
    )
    render_themed_table(styled)

    # QoQ summary pills
    if not sub_q.empty:
        pills = ""
        for _, qrow in sub_q.iterrows():
            qoq = qrow['MoM_Pct']
            yoy = qrow['YoY_Pct']
            color = MONTHLY_COLOR if (yoy or 0) >= 4 else ("#F59E0B" if (yoy or 0) >= 2 else "#16A34A")
            qoq_str = f"{qoq:+.2f}%" if qoq is not None else "—"
            yoy_str = f"{yoy:+.1f}% YoY" if yoy is not None else "—"
            pills += f'<span style="display:inline-block;background:{PILL_BG};border:1px solid {PILL_BORDER};color:{PILL_FG};border-radius:8px;padding:8px 14px;margin:4px;font-family:IBM Plex Mono,monospace;font-size:0.82rem;"><strong style="color:{color}">{qrow["Period"]}</strong><br>QoQ {qoq_str} &nbsp;|&nbsp; {yoy_str}</span>'
        st.markdown(f'<div style="margin-top:12px;">{pills}</div>', unsafe_allow_html=True)


# ─── SHARED RENDER HELPERS ────────────────────────────────────────────────────

def _render_table(sub, city_df, start_period, end_period, sv, frequency, compact=False):
    pop_label = "QoQ (%)" if frequency == "Quarterly" else "MoM (%)"
    sub_display = sub.copy()
    sub_display['vs_Start_%'] = ((sub_display['CPI_Index'] - sv) / sv * 100).round(2)
    display_df = sub_display[['Period', 'CPI_Index', 'MoM_Pct', 'YoY_Pct', 'vs_Start_%']].copy()
    display_df.columns = ['Period', 'CPI Index', pop_label, 'YoY (%)', 'vs Start (%)']

    def color_pct(val):
        if pd.isna(val) or not isinstance(val, (int, float)): return ''
        if val > 0: return f'color: {POS_COLOR}; font-weight: 600'
        if val < 0: return f'color: {NEG_COLOR}; font-weight: 600'
        return ''

    styled = (display_df.style
        .format({'CPI Index': '{:.2f}',
                 pop_label: lambda x: f'{x:+.1f}%' if pd.notna(x) else '—',
                 'YoY (%)': lambda x: f'{x:+.1f}%' if pd.notna(x) else '—',
                 'vs Start (%)': '{:+.2f}%'})
        .map(color_pct, subset=[pop_label, 'YoY (%)', 'vs Start (%)'])
    )
    if not compact:
        st.markdown('<p class="sec-title">Detailed Results Table</p>', unsafe_allow_html=True)
    render_themed_table(styled)


def _render_snapshot(df, city):
    with st.expander("📊 All Regions Snapshot — Latest Period"):
        latest_row = df.iloc[-1]
        latest_period = latest_row['Period']
        snap_data = []
        for c in CITIES:
            idx_val = latest_row[f'{c}_Index']
            yoy_val = latest_row[f'{c}_YoY']
            mom_val = latest_row[f'{c}_MoM']
            signal = "🔴" if (yoy_val or 0) >= 4 else ("🟡" if (yoy_val or 0) >= 2 else "🟢")
            snap_data.append({
                'Region': c,
                'CPI Index': f"{idx_val:.2f}" if idx_val else '—',
                'YoY % (Official)': f"{yoy_val:+.1f}%" if yoy_val else '—',
                'MoM %': f"{mom_val:+.1f}%" if mom_val else '—',
                'Signal': signal,
            })
        st.caption(f"Reference period: **{latest_period}** — Source: ABS 6401.0 (live)")
        render_themed_table(pd.DataFrame(snap_data))


def _render_exports(city_df, city, start_period, end_period, calc, frequency):
    st.markdown('<p class="sec-title">Export</p>', unsafe_allow_html=True)
    sub = filter_city_df(city_df, start_period, end_period)
    pop_label = "QoQ (%)" if frequency == "Quarterly" else "MoM (%)"
    sub_display = sub[['Period', 'CPI_Index', 'MoM_Pct', 'YoY_Pct']].copy()
    sub_display.columns = ['Period', 'CPI Index', pop_label, 'YoY (%)']
    ec1, ec2, ec3 = st.columns([1, 1, 3])
    with ec1:
        csv_out = sub_display.to_csv(index=False).encode("utf-8")
        st.download_button("⬇ Download CSV", data=csv_out,
                           file_name=f"CPI_{city}_{start_period}_{end_period}.csv",
                           mime="text/csv")
    with ec2:
        xlsx_out = build_excel_report(city_df, city, start_period, end_period, calc, frequency)
        st.download_button("⬇ Download Excel Report", data=xlsx_out,
                           file_name=f"CPI_Report_{city}_{start_period}_{end_period}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _render_footer(df, frequency_label):
    st.markdown("---")
    latest_period = df['Period'].iloc[-1]
    st.markdown(f"""
    <div style="font-size:0.75rem;color:#bbb;text-align:center;padding:6px 0;">
        Australia Post · CPI Calculator v3.1 · Data & Pricing Executive Tool ·
        Source: <a href="https://www.abs.gov.au/statistics/economy/price-indexes-and-inflation/consumer-price-index-australia/latest-release" target="_blank" style="color:#FF5A73;">ABS 6401.0 ({frequency_label}, {latest_period})</a> ·
        Base Sep-2025 = 100 · 🟢 Live Data
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
